"""将 pose JSON 中的 COCO-17 骨架导出为 xlsx；碰撞/告警帧额外写入事件表。"""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any

import cv2

from model_assets import COCO17_KEYPOINT_NAMES
from event_engine.annotation_boxes import (
    build_scaled_boxes,
    flatten_annotation_boxes,
    load_scaled_boxes,
)
from event_engine.box_identity import box_collision_token
from event_engine.collision_methods import build_collision_params, create_collision_processor

# 关键点中文名（与 COCO17 顺序一致，便于表头阅读）
COCO17_KEYPOINT_LABELS_ZH = (
    "鼻子",
    "左眼",
    "右眼",
    "左耳",
    "右耳",
    "左肩",
    "右肩",
    "左肘",
    "右肘",
    "左腕",
    "右腕",
    "左髋",
    "右髋",
    "左膝",
    "右膝",
    "左踝",
    "右踝",
)


def _frame_collision_idx(frame: dict[str, Any]) -> int:
    return int(frame.get("source_frame_idx") or frame.get("frame_idx") or 0)


def _infer_size_from_pose(pose_data: dict[str, Any]) -> tuple[int, int]:
    infer_w = int(pose_data.get("infer_width") or 0)
    infer_h = int(pose_data.get("infer_height") or 0)
    if infer_w > 0 and infer_h > 0:
        return infer_w, infer_h
    frames = pose_data.get("frames") or []
    if frames and isinstance(frames[0], dict):
        fw = int(frames[0].get("infer_width") or 0)
        fh = int(frames[0].get("infer_height") or 0)
        if fw > 0 and fh > 0:
            return fw, fh
    ann = pose_data.get("annotation") or {}
    if isinstance(ann, dict):
        sz = ann.get("annotation_size") or {}
        if isinstance(sz, dict):
            w = int(sz.get("width") or 0)
            h = int(sz.get("height") or 0)
            if w > 0 and h > 0:
                return w, h
    return 640, 480


def _scaled_boxes_for_pose(
    pose_data: dict[str, Any],
    annotation_path: Path | None,
    infer_w: int,
    infer_h: int,
) -> list[dict[str, Any]]:
    if annotation_path and annotation_path.is_file():
        return load_scaled_boxes(annotation_path, infer_w, infer_h)

    ann = pose_data.get("annotation")
    if not isinstance(ann, dict):
        return []

    raw_boxes = ann.get("boxes")
    if not isinstance(raw_boxes, list) or not raw_boxes:
        raw_boxes = flatten_annotation_boxes(ann)
    if not raw_boxes:
        return []

    ann_sz = ann.get("annotation_size") or ann.get("source_annotation_size") or {}
    ann_w = ann_h = None
    if isinstance(ann_sz, dict):
        try:
            if ann_sz.get("width") is not None:
                ann_w = float(ann_sz["width"])
            if ann_sz.get("height") is not None:
                ann_h = float(ann_sz["height"])
        except (TypeError, ValueError):
            ann_w = ann_h = None
    return build_scaled_boxes(raw_boxes, ann_w, ann_h, infer_w, infer_h)


def _frames_need_collision_recompute(frames: list[dict[str, Any]]) -> bool:
    """仅当帧内完全无碰撞字段、或字段恒为空且未在采集时启用碰撞时才离线重算。"""
    if not frames:
        return False
    for fr in frames:
        if not isinstance(fr, dict):
            continue
        if fr.get("collisions") or fr.get("alarm_collisions"):
            return False
    return True


_EVENT_TYPE_ZH = {"alarm": "告警", "collision": "碰撞"}


def _verified_lookup_from_review(review: dict[str, Any] | None) -> set[tuple[str, int, str]]:
    """人工复核：{(事件类型中文, 帧序号, 货框标识), ...}。"""
    out: set[tuple[str, int, str]] = set()
    if not isinstance(review, dict):
        return out
    for item in review.get("verified_true") or []:
        if not isinstance(item, dict):
            continue
        et = _EVENT_TYPE_ZH.get(str(item.get("event_type") or "").strip(), "")
        if not et:
            continue
        try:
            fi = int(item.get("frame_idx") or 0)
        except (TypeError, ValueError):
            continue
        for token in item.get("box_tokens") or []:
            t = str(token).strip()
            if t:
                out.add((et, fi, t))
    return out


def _human_verified_label(
    verified_lookup: set[tuple[str, int, str]],
    *,
    event_type_zh: str,
    frame_idx: Any,
    token: str,
) -> str:
    """人工标真：是 / 未复核（仅标记真实碰撞，未标默认为未复核）。"""
    if not verified_lookup:
        return "未复核"
    try:
        fi = int(frame_idx or 0)
    except (TypeError, ValueError):
        return "未复核"
    t = str(token or "").strip()
    if not t:
        return "未复核"
    return "是" if (event_type_zh, fi, t) in verified_lookup else "未复核"


def _excel_cell(value: Any) -> Any:
    """Parquet/numpy 标量转为 openpyxl 可写入类型。"""
    if value is None:
        return None
    try:
        import numpy as np

        if isinstance(value, np.generic):
            return value.item()
    except ImportError:
        pass
    if isinstance(value, (list, tuple)):
        return ";".join(str(v) for v in value if v is not None and str(v).strip())
    return value


def enrich_frames_with_collision(
    pose_data: dict[str, Any],
    *,
    annotation_path: Path | None = None,
    alarm_min_consecutive_frames: int = 3,
    alarm_cooldown_frames: int = 12,
) -> dict[str, Any]:
    """若帧内无碰撞字段且存在标注，离线重算 collisions / alarm_collisions。"""
    frames = pose_data.get("frames")
    if not isinstance(frames, list) or not frames:
        return pose_data

    if not _frames_need_collision_recompute(frames):
        return pose_data

    infer_w, infer_h = _infer_size_from_pose(pose_data)
    scaled_boxes = _scaled_boxes_for_pose(pose_data, annotation_path, infer_w, infer_h)
    if not scaled_boxes:
        return pose_data

    fps = float(pose_data.get("fps") or 15.0)
    if fps <= 0:
        fps = 15.0

    collision_cfg = pose_data.get("collision") or {}
    if isinstance(collision_cfg, dict):
        params = build_collision_params(
            collision_cfg.get("method") or "wrist_point",
            collision_cfg,
            alarm_min_consecutive_frames=int(
                collision_cfg.get("alarm_min_consecutive_frames") or alarm_min_consecutive_frames
            ),
            alarm_cooldown_frames=int(collision_cfg.get("alarm_cooldown_frames") or alarm_cooldown_frames),
        )
    else:
        params = build_collision_params(
            "wrist_point",
            alarm_min_consecutive_frames=alarm_min_consecutive_frames,
            alarm_cooldown_frames=alarm_cooldown_frames,
        )

    processor = create_collision_processor(
        scaled_boxes,
        method=params.get("method"),
        params=params,
        video_fps=fps,
    )

    out_frames: list[dict[str, Any]] = []
    for fr in frames:
        if not isinstance(fr, dict):
            continue
        frame_copy = dict(fr)
        idx = _frame_collision_idx(fr)
        event = processor.process({"frame_idx": idx, "persons": fr.get("persons") or []})
        frame_copy["persons"] = event.get("skeletons") or fr.get("persons") or []
        frame_copy["collisions"] = event.get("collisions") or []
        frame_copy["alarm_collisions"] = event.get("alarm_collisions") or []
        out_frames.append(frame_copy)

    enriched = dict(pose_data)
    enriched["frames"] = out_frames
    enriched.setdefault("collision", {})
    if isinstance(enriched["collision"], dict):
        enriched["collision"]["enabled"] = True
        enriched["collision"]["recomputed_on_export"] = True
    return enriched


def _person_wrist_hits(
    person: dict[str, Any],
    boxes: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """检测人员手腕进入的货框，返回 [{wrist, token}, ...]。"""
    hits: list[dict[str, Any]] = []
    keypoints = person.get("keypoints") or []
    wrist_names = ((9, "left_wrist"), (10, "right_wrist"))
    for kpt_idx, wrist_name in wrist_names:
        if len(keypoints) <= kpt_idx:
            continue
        kp = keypoints[kpt_idx]
        if not isinstance(kp, (list, tuple)) or len(kp) < 3:
            continue
        if float(kp[2]) <= 0.3:
            continue
        wx, wy = float(kp[0]), float(kp[1])
        for box in boxes:
            contour = box.get("orig_contour")
            if contour is None:
                continue
            if cv2.pointPolygonTest(contour, (wx, wy), False) >= 0:
                token = box_collision_token(box)
                if token:
                    hits.append({"wrist": wrist_name, "token": token, "x": wx, "y": wy})
                break
    return hits


def _keypoint_headers(names: list[str]) -> list[str]:
    cols: list[str] = []
    for i, name in enumerate(names):
        zh = COCO17_KEYPOINT_LABELS_ZH[i] if i < len(COCO17_KEYPOINT_LABELS_ZH) else name
        cols.extend([f"{zh}({name})_x", f"{zh}({name})_y", f"{zh}({name})_score"])
    return cols


def _append_person_row(
    row: list[Any],
    *,
    frame: dict[str, Any],
    person: dict[str, Any],
    keypoint_names: list[str],
    collisions: list[str],
    alarms: list[str],
) -> list[Any]:
    bbox = person.get("bbox") or [None, None, None, None]
    if not isinstance(bbox, (list, tuple)):
        bbox = [None, None, None, None]
    kpts = person.get("keypoints") or []
    out = list(row)
    out.extend(
        [
            _excel_cell(frame.get("frame_idx")),
            _excel_cell(frame.get("source_frame_idx")),
            _excel_cell(frame.get("timestamp_sec")),
            _excel_cell(person.get("person_id")),
            _excel_cell(person.get("person_track_id")),
            _excel_cell(bbox[0] if len(bbox) > 0 else None),
            _excel_cell(bbox[1] if len(bbox) > 1 else None),
            _excel_cell(bbox[2] if len(bbox) > 2 else None),
            _excel_cell(bbox[3] if len(bbox) > 3 else None),
        ]
    )
    for i in range(len(keypoint_names)):
        if i < len(kpts) and isinstance(kpts[i], (list, tuple)) and len(kpts[i]) >= 2:
            kp = kpts[i]
            out.extend([float(kp[0]), float(kp[1]), float(kp[2]) if len(kp) > 2 else None])
        else:
            out.extend([None, None, None])
    out.extend(
        [
            "是" if collisions else "否",
            "是" if alarms else "否",
            ";".join(str(t) for t in collisions if t is not None and str(t).strip()),
            ";".join(str(t) for t in alarms if t is not None and str(t).strip()),
        ]
    )
    return out


def export_pose_to_xlsx_bytes(
    pose_data: dict[str, Any],
    *,
    annotation_path: Path | None = None,
    alarm_min_consecutive_frames: int = 3,
    alarm_cooldown_frames: int = 12,
    event_review: dict[str, Any] | None = None,
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，请执行: pip install openpyxl") from exc

    data = enrich_frames_with_collision(
        pose_data,
        annotation_path=annotation_path,
        alarm_min_consecutive_frames=alarm_min_consecutive_frames,
        alarm_cooldown_frames=alarm_cooldown_frames,
    )

    keypoint_names = list(data.get("keypoint_names") or COCO17_KEYPOINT_NAMES)
    frames = data.get("frames") or []

    infer_w, infer_h = _infer_size_from_pose(data)
    scaled_boxes = _scaled_boxes_for_pose(data, annotation_path, infer_w, infer_h)
    review = event_review if isinstance(event_review, dict) else data.get("event_review")
    verified_lookup = _verified_lookup_from_review(review if isinstance(review, dict) else None)

    base_headers = [
        "帧序号",
        "源视频帧序号",
        "时间戳(秒)",
        "人员ID",
        "跟踪ID",
        "bbox_x1",
        "bbox_y1",
        "bbox_x2",
        "bbox_y2",
    ]
    kpt_headers = _keypoint_headers(keypoint_names)
    tail_headers = ["本帧有碰撞", "本帧有告警", "碰撞货框", "告警货框"]
    skeleton_headers = base_headers + kpt_headers + tail_headers

    event_headers = [
        "事件类型",
        "帧序号",
        "源视频帧序号",
        "时间戳(秒)",
        "人员ID",
        "跟踪ID",
        "触发手腕",
        "货框标识",
        "人工标真",
        "手腕_x",
        "手腕_y",
    ] + kpt_headers

    wb = Workbook()
    ws_sk = wb.active
    ws_sk.title = "骨架数据"
    ws_sk.append(skeleton_headers)

    ws_ev = wb.create_sheet("碰撞告警事件")
    ws_ev.append(event_headers)

    seen_event_keys: set[tuple[Any, ...]] = set()

    for frame in frames:
        if not isinstance(frame, dict):
            continue
        persons = frame.get("persons") or []
        frame_collisions = list(frame.get("collisions") or [])
        frame_alarms = list(frame.get("alarm_collisions") or [])

        if not persons:
            ws_sk.append(
                _append_person_row(
                    [],
                    frame=frame,
                    person={},
                    keypoint_names=keypoint_names,
                    collisions=frame_collisions,
                    alarms=frame_alarms,
                )
            )

        for person in persons:
            if not isinstance(person, dict):
                continue
            ws_sk.append(
                _append_person_row(
                    [],
                    frame=frame,
                    person=person,
                    keypoint_names=keypoint_names,
                    collisions=frame_collisions,
                    alarms=frame_alarms,
                )
            )

            if not scaled_boxes:
                continue

            hits = _person_wrist_hits(person, scaled_boxes)
            for hit in hits:
                token = hit["token"]
                is_alarm = token in frame_alarms
                is_collision = token in frame_collisions or is_alarm
                if not is_collision and not is_alarm:
                    continue

                for event_type, active in (("告警", is_alarm), ("碰撞", is_collision and not is_alarm)):
                    if not active:
                        continue
                    ev_key = (
                        event_type,
                        frame.get("frame_idx"),
                        person.get("person_id"),
                        token,
                        hit.get("wrist"),
                    )
                    if ev_key in seen_event_keys:
                        continue
                    seen_event_keys.add(ev_key)

                    ev_row: list[Any] = [
                        event_type,
                        _excel_cell(frame.get("frame_idx")),
                        _excel_cell(frame.get("source_frame_idx")),
                        _excel_cell(frame.get("timestamp_sec")),
                        _excel_cell(person.get("person_id")),
                        _excel_cell(person.get("person_track_id")),
                        hit.get("wrist"),
                        token,
                        _human_verified_label(
                            verified_lookup,
                            event_type_zh=event_type,
                            frame_idx=frame.get("frame_idx"),
                            token=token,
                        ),
                        _excel_cell(hit.get("x")),
                        _excel_cell(hit.get("y")),
                    ]
                    kpts = person.get("keypoints") or []
                    for i in range(len(keypoint_names)):
                        if i < len(kpts) and isinstance(kpts[i], (list, tuple)) and len(kpts[i]) >= 2:
                            kp = kpts[i]
                            ev_row.extend([float(kp[0]), float(kp[1]), float(kp[2]) if len(kp) > 2 else None])
                        else:
                            ev_row.extend([None, None, None])
                    ws_ev.append(ev_row)

        # 帧级碰撞但未能关联到具体人员时，仍写入事件表
        if scaled_boxes and frame_collisions and not persons:
            for token in frame_collisions:
                ev_key = ("碰撞", frame.get("frame_idx"), None, token, None)
                if ev_key in seen_event_keys:
                    continue
                seen_event_keys.add(ev_key)
                ev_type_zh = "告警" if token in frame_alarms else "碰撞"
                ws_ev.append(
                    [
                        ev_type_zh,
                        _excel_cell(frame.get("frame_idx")),
                        _excel_cell(frame.get("source_frame_idx")),
                        _excel_cell(frame.get("timestamp_sec")),
                        None,
                        None,
                        "",
                        token,
                        _human_verified_label(
                            verified_lookup,
                            event_type_zh=ev_type_zh,
                            frame_idx=frame.get("frame_idx"),
                            token=token,
                        ),
                        None,
                        None,
                    ]
                    + [None] * (3 * len(keypoint_names))
                )

    if isinstance(review, dict) and review.get("verified_true"):
        ws_review = wb.create_sheet("人工复核")
        ws_review.append(["事件类型", "帧序号", "源视频帧序号", "货框标识", "复核结果", "更新时间"])
        updated_at = str(review.get("updated_at") or "")
        for item in review.get("verified_true") or []:
            if not isinstance(item, dict):
                continue
            et = _EVENT_TYPE_ZH.get(str(item.get("event_type") or "").strip(), str(item.get("event_type") or ""))
            tokens = [str(t).strip() for t in (item.get("box_tokens") or []) if str(t).strip()]
            for token in tokens or [""]:
                ws_review.append(
                    [
                        et,
                        _excel_cell(item.get("frame_idx")),
                        _excel_cell(item.get("source_frame_idx")),
                        token,
                        "是",
                        updated_at,
                    ]
                )

    ws_info = wb.create_sheet("说明")
    ws_info.append(["字段", "说明"])
    ws_info.append(["骨架数据", "每帧每位人员一行，含 COCO-17 共 17 个关键点 x/y/score"])
    ws_info.append(["碰撞告警事件", "手腕进入货框（碰撞）或连续帧触发告警时额外记录，含完整骨架"])
    ws_info.append(["人工标真", "回放复核后写入：是=人工确认为真实碰撞/告警；未复核=尚未标真"])
    ws_info.append(["人工复核", "汇总已标真事件（与 event_review.json 一致）"])
    verified_n = len(review.get("verified_true") or []) if isinstance(review, dict) else 0
    ws_info.append(["已标真事件数", verified_n])
    ws_info.append(["源视频", data.get("source_video") or ""])
    ws_info.append(["模型", data.get("model") or ""])
    ws_info.append(["总帧数", data.get("frame_count") or len(frames)])
    ws_info.append(["关键点名称", ", ".join(keypoint_names)])

    # 略加宽首行
    for ws in (ws_sk, ws_ev):
        for col_idx in range(1, min(len(skeleton_headers), 12) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pose_json_file_to_xlsx(
    pose_json_path: Path,
    output_path: Path,
    *,
    annotation_path: Path | None = None,
    alarm_min_consecutive_frames: int = 3,
    alarm_cooldown_frames: int = 12,
    event_review_path: Path | None = None,
) -> Path:
    with open(pose_json_path, encoding="utf-8") as f:
        pose_data = json.load(f)
    review: dict[str, Any] | None = None
    review_path = event_review_path or pose_json_path.with_suffix(".event_review.json")
    if review_path.is_file():
        try:
            with open(review_path, encoding="utf-8") as f:
                raw = json.load(f)
            if isinstance(raw, dict):
                review = raw
        except (OSError, json.JSONDecodeError):
            review = None
    blob = export_pose_to_xlsx_bytes(
        pose_data,
        annotation_path=annotation_path,
        alarm_min_consecutive_frames=alarm_min_consecutive_frames,
        alarm_cooldown_frames=alarm_cooldown_frames,
        event_review=review,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(blob)
    return output_path.resolve()
